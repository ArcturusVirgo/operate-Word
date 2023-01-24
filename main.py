import csv
import os
from pprint import pprint

from docx import Document
from openpyxl import load_workbook
import pandas as pd

data_all = {}
repeat = {}
files = os.listdir('./docx')
text = []

wb = load_workbook('发展对象.xlsx')
sheet = wb['Sheet1']
rows = sheet.rows

for file in files:
    data_person = {}
    path = f'./docx/{file}'
    scored = file.strip('.docx')
    document = Document(path)  # 读入文件
    tables = document.tables  # 获取文件中的表格集
    table = tables[0]  # 获取文件中的第一个表格
    for i in range(3, len(table.rows) - 1):  # 从表格第二行开始循环读取表格数据
        num = table.cell(i, 0).text
        name = table.cell(i, 1).text
        class_ = table.cell(i, 2).text
        score_1 = table.cell(i, 3).text
        score_2 = table.cell(i, 4).text
        score_3 = table.cell(i, 5).text

        data_person[name] = data_person.get(name, []) + [score_1, score_2, score_3]

    data_all[scored] = data_person
pprint(data_all)
